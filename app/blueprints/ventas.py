# -*- coding: utf-8 -*-
"""
KATITA-POS - Blueprint de Ventas
=================================
Endpoints para procesar ventas con sistema FIFO automatico.

Este modulo maneja el CORE del negocio:
- Procesamiento de ventas con descuento automatico FIFO de lotes
- Trazabilidad completa: cada venta sabe de que lote salio
- Metodos de pago peruanos (efectivo, yape, plin, transferencia)
- Calculo automatico de cambio y totales
- Reportes y estadisticas de ventas

El sistema FIFO es CRITICO:
- Primero que vence, primero que sale
- Si un lote no alcanza, usa el siguiente automaticamente
- Cada descuento genera MovimientoStock para auditoria
- Atomicidad total: todo o nada (rollback en errores)
"""

from flask import Blueprint, request, g, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity, get_jwt
from sqlalchemy.exc import IntegrityError
from sqlalchemy import func, and_
from decimal import Decimal
from datetime import datetime, timezone, date, timedelta
from io import BytesIO

# Zona horaria de Perú (UTC-5)
PERU_TZ = timezone(timedelta(hours=-5))
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app import db
from app.models.venta import Venta
from app.models.detalle_venta import DetalleVenta
from app.models.product import Product
from app.models.lote import Lote
from app.models.movimiento_stock import MovimientoStock
from app.models.user import User
from app.models.cuadro_caja import CuadroCaja
from app.utils.responses import (
    success_response, error_response, created_response,
    not_found_response, validation_error_response
)
from app.decorators.auth_decorators import login_required, role_required

# Crear Blueprint con prefijo /api/ventas
ventas_bp = Blueprint('ventas', __name__, url_prefix='/api/ventas')


# ==================================================================================
# ENDPOINT 1: POST /api/ventas - Procesar venta con FIFO (ENDPOINT MAS CRITICO)
# ==================================================================================

@ventas_bp.route('', methods=['POST'])
@jwt_required()
def procesar_venta():
    """
    Procesar una nueva venta con descuento automatico FIFO de lotes

    Este es el endpoint MAS CRITICO del sistema. Cuando el vendedor procesa
    una venta, el sistema automaticamente:

    1. Valida que haya stock suficiente para todos los items
    2. Descuenta del lote que vence primero (FIFO)
    3. Si un lote no alcanza, usa el siguiente lote automaticamente
    4. Registra trazabilidad completa (cada item sabe de que lote salio)
    5. Actualiza stock de productos y lotes
    6. Genera MovimientoStock por cada descuento
    7. Calcula totales y cambio
    8. Genera numero de venta unico

    ATOMICIDAD: Si algo falla, hace rollback completo (todo o nada)

    Body JSON:
        {
            "vendedor_id": 1,
            "items": [
                {
                    "producto_id": 1,
                    "cantidad": 10,
                    "precio_unitario": 3.50
                },
                {
                    "producto_id": 2,
                    "cantidad": 5,
                    "precio_unitario": 2.00
                }
            ],
            "metodo_pago": "efectivo",
            "monto_recibido": 60.00,
            "cliente_nombre": "Juan Perez",
            "cliente_dni": "12345678",
            "descuento": 2.50,
            "notas": "Cliente frecuente"
        }

    Validaciones:
        - vendedor_id: Debe existir y estar activo
        - items: No puede estar vacio
        - metodo_pago: Debe ser 'efectivo', 'yape', 'plin' o 'transferencia'
        - monto_recibido: Requerido SOLO para efectivo, debe ser >= total
        - Para pagos digitales: monto_recibido debe ser null
        - Cada item: producto existe, activo, cantidad > 0, stock suficiente

    Returns:
        201: Venta procesada exitosamente
        400: Error de validacion
        404: Producto/vendedor no encontrado
        422: Stock insuficiente
        500: Error interno

    Ejemplo de respuesta exitosa:
        {
            "success": true,
            "message": "Venta procesada exitosamente",
            "data": {
                "venta": {
                    "id": 1,
                    "numero_venta": "V-20251104-0001",
                    "subtotal": 45.00,
                    "descuento": 2.50,
                    "total": 42.50,
                    "metodo_pago": "efectivo",
                    "monto_recibido": 60.00,
                    "cambio": 17.50,
                    "estado": "completada"
                },
                "detalles": [
                    {
                        "producto": "Coca Cola 500ml",
                        "cantidad": 10,
                        "precio_unitario": 3.50,
                        "subtotal": 35.00,
                        "lote": "LOTE-20251104-002"
                    }
                ]
            }
        }
    """
    try:
        from flask import current_app
        current_app.logger.info("[VENTAS] Iniciando creacion de venta...")

        print('\n' + '='*70)
        print('📥 RECIBIDA PETICIÓN: POST /api/ventas')
        print('='*70)

        data = request.json
        current_app.logger.info(f"[VENTAS] Datos recibidos: {data}")

        print('📦 DATOS RECIBIDOS (RAW):')
        import json
        print(json.dumps(data, indent=2, ensure_ascii=False))

        print(f'\n🔑 Campos recibidos: {list(data.keys())}')
        print(f'📊 Tiene "items": {"items" in data}')
        print(f'📊 Tiene "detalles": {"detalles" in data}')

        # ========== VALIDACIONES INICIALES ==========
        errores = {}

        # Obtener vendedor_id del usuario autenticado (del token JWT)
        current_user_id_str = get_jwt_identity()
        vendedor_id = int(current_user_id_str) if current_user_id_str else None
        print(f'👤 Vendedor ID (del token JWT): {vendedor_id}')

        # Validar campos requeridos (aceptar tanto "items" como "detalles")
        items = data.get('items') or data.get('detalles')
        if not items or len(items) == 0:
            print('❌ ERROR: Falta campo "items"/"detalles" o está vacío')
            errores['items'] = 'Debe incluir al menos un item'

        if not data.get('metodo_pago'):
            print('❌ ERROR: Falta campo "metodo_pago"')
            errores['metodo_pago'] = 'El metodo de pago es requerido'

        if errores:
            print(f'❌ ERRORES DE VALIDACIÓN INICIALES: {errores}')
            print('='*70 + '\n')
            return validation_error_response(errores)

        # Validar vendedor existe y esta activo
        vendedor = User.query.get(vendedor_id)
        if not vendedor:
            return not_found_response(f"Vendedor con ID {vendedor_id} no encontrado")

        if not vendedor.activo:
            return error_response('El vendedor no esta activo', status_code=400)

        # Validar metodo de pago
        metodos_validos = ['efectivo', 'yape', 'plin', 'transferencia']
        metodo_pago = data['metodo_pago'].lower()
        if metodo_pago not in metodos_validos:
            return error_response(
                f"Metodo de pago invalido. Use: {', '.join(metodos_validos)}",
                status_code=400
            )

        # Validar monto_recibido segun metodo de pago
        monto_recibido = data.get('monto_recibido')
        if metodo_pago == 'efectivo':
            if monto_recibido is None:
                errores['monto_recibido'] = 'El monto recibido es requerido para pago en efectivo'
            else:
                try:
                    monto_recibido = Decimal(str(monto_recibido))
                    if monto_recibido <= 0:
                        errores['monto_recibido'] = 'Debe ser mayor a 0'
                except (ValueError, TypeError):
                    errores['monto_recibido'] = 'Debe ser un numero valido'
        else:
            # Para pagos digitales, monto_recibido debe ser None
            monto_recibido = None

        if errores:
            return validation_error_response(errores)

        # ========== VALIDAR ITEMS Y VERIFICAR STOCK DISPONIBLE ==========
        items_validados = []

        for idx, item in enumerate(items):
            item_errores = {}

            # Validar campos del item
            if 'producto_id' not in item:
                item_errores['producto_id'] = 'El producto_id es requerido'
            if 'cantidad' not in item:
                item_errores['cantidad'] = 'La cantidad es requerida'
            if 'precio_unitario' not in item:
                item_errores['precio_unitario'] = 'El precio_unitario es requerido'

            if item_errores:
                errores[f'item_{idx}'] = item_errores
                continue

            # Validar producto existe y esta activo
            producto = Product.query.get(item['producto_id'])
            if not producto:
                errores[f'item_{idx}'] = f"Producto con ID {item['producto_id']} no encontrado"
                continue

            if not producto.activo:
                errores[f'item_{idx}'] = f"El producto {producto.nombre} no esta activo"
                continue

            # Validar cantidad
            try:
                cantidad = int(item['cantidad'])
                if cantidad <= 0:
                    item_errores['cantidad'] = 'Debe ser mayor a 0'
            except (ValueError, TypeError):
                item_errores['cantidad'] = 'Debe ser un numero entero'
                cantidad = 0

            # Validar precio unitario
            try:
                precio_unitario = Decimal(str(item['precio_unitario']))
                if precio_unitario <= 0:
                    item_errores['precio_unitario'] = 'Debe ser mayor a 0'
            except (ValueError, TypeError):
                item_errores['precio_unitario'] = 'Debe ser un numero decimal valido'
                precio_unitario = Decimal('0')

            if item_errores:
                errores[f'item_{idx}'] = item_errores
                continue

            # Validar stock disponible (suma de lotes FIFO)
            lotes_disponibles = Lote.lotes_fifo(producto.id).all()
            stock_total_disponible = sum(lote.cantidad_actual for lote in lotes_disponibles)

            if stock_total_disponible < cantidad:
                errores[f'item_{idx}'] = (
                    f"Stock insuficiente para {producto.nombre}. "
                    f"Solicitado: {cantidad}, Disponible: {stock_total_disponible}"
                )
                continue

            # Item validado correctamente
            items_validados.append({
                'producto': producto,
                'lotes_fifo': lotes_disponibles,
                'cantidad': cantidad,
                'precio_unitario': precio_unitario
            })

        if errores:
            return validation_error_response(errores)

        # ========== BUSCAR TURNO ABIERTO DEL VENDEDOR ==========
        turno_abierto = CuadroCaja.query.filter_by(
            vendedor_id=vendedor_id,
            estado='abierto'
        ).first()

        # ========== CREAR VENTA MAESTRA ==========
        descuento = Decimal(str(data.get('descuento', 0)))

        # Calcular subtotal primero para inicializar la venta correctamente
        subtotal_temp = sum(
            Decimal(str(item['precio_unitario'])) * item['cantidad']
            for item in items_validados
        )
        total_temp = (subtotal_temp - descuento).quantize(Decimal('0.01'))

        nueva_venta = Venta(
            vendedor_id=vendedor_id,
            cuadro_caja_id=turno_abierto.id if turno_abierto else None,
            metodo_pago=metodo_pago,
            monto_recibido=monto_recibido,
            cliente_nombre=data.get('cliente_nombre', ''),
            cliente_dni=data.get('cliente_dni', ''),
            descuento=descuento,
            notas=data.get('notas', ''),
            estado='completada',
            created_offline=False,
            synced=False,
            subtotal=subtotal_temp,
            total=total_temp
        )

        # Generar numero de venta unico
        nueva_venta.generar_numero_venta()

        db.session.add(nueva_venta)
        db.session.flush()  # Para obtener el ID de la venta

        # ========== PROCESAR CADA ITEM CON FIFO ==========
        detalles_creados = []

        for item_data in items_validados:
            producto = item_data['producto']
            lotes_fifo = item_data['lotes_fifo']
            cantidad_solicitada = item_data['cantidad']
            precio_unitario = item_data['precio_unitario']

            cantidad_restante = cantidad_solicitada

            # Descontar usando FIFO: primero del lote que vence antes
            for lote in lotes_fifo:
                if cantidad_restante <= 0:
                    break  # Ya se proceso toda la cantidad

                # Calcular cuanto descontar de este lote
                cantidad_a_descontar = min(cantidad_restante, lote.cantidad_actual)

                # Guardar stock anterior para MovimientoStock
                stock_anterior = producto.stock_total

                # Descontar del lote (usa el metodo del modelo)
                lote.descontar_stock(cantidad_a_descontar)

                # Actualizar stock del producto
                producto.stock_total -= cantidad_a_descontar
                stock_nuevo = producto.stock_total

                # Crear DetalleVenta con trazabilidad del lote
                detalle = DetalleVenta(
                    venta_id=nueva_venta.id,
                    producto_id=producto.id,
                    lote_id=lote.id,
                    cantidad=cantidad_a_descontar,
                    precio_unitario=precio_unitario,
                    precio_compra=lote.precio_compra_lote
                )

                # Calcular subtotales (ganancia, etc)
                detalle.calcular_subtotales()

                db.session.add(detalle)
                detalles_creados.append(detalle)

                # Crear MovimientoStock para auditoria
                movimiento = MovimientoStock(
                    tipo='venta',
                    producto_id=producto.id,
                    lote_id=lote.id,
                    usuario_id=vendedor_id,
                    venta_id=nueva_venta.id,
                    cantidad=-cantidad_a_descontar,  # Negativo porque es salida
                    stock_anterior=stock_anterior,
                    stock_nuevo=stock_nuevo,
                    motivo=f'Venta {nueva_venta.numero_venta}',
                    referencia=nueva_venta.numero_venta
                )

                db.session.add(movimiento)

                # Restar lo que ya se desconto
                cantidad_restante -= cantidad_a_descontar

            # Verificacion de seguridad: deberia haber procesado todo
            if cantidad_restante > 0:
                raise ValueError(
                    f"Error CRITICO: No se pudo procesar {cantidad_restante} unidades "
                    f"de {producto.nombre}. Esto NO deberia pasar."
                )

        # Los totales ya fueron calculados antes del flush, no es necesario recalcular

        # ========== VALIDAR MONTO RECIBIDO Y CALCULAR CAMBIO ==========
        if metodo_pago == 'efectivo':
            if monto_recibido < nueva_venta.total:
                raise ValueError(
                    f"Monto recibido insuficiente. "
                    f"Total: {nueva_venta.total}, Recibido: {monto_recibido}"
                )
            nueva_venta.cambio = monto_recibido - nueva_venta.total
        else:
            # Para pagos digitales, cambio siempre es 0
            nueva_venta.cambio = Decimal('0')

        # ========== REGISTRAR EN CUADRO DE CAJA ==========
        # Si el vendedor tiene un turno abierto, registrar la venta automáticamente
        turno_activo = CuadroCaja.turno_abierto_vendedor(vendedor_id)
        if turno_activo:
            try:
                turno_activo.registrar_venta(nueva_venta)
                current_app.logger.info(f"[VENTAS] Venta registrada en cuadro de caja {turno_activo.numero_turno}")
            except Exception as e:
                current_app.logger.warning(f"[VENTAS] No se pudo registrar en cuadro de caja: {str(e)}")
                # No lanzar error, solo advertencia - la venta debe procesarse igual

        # ========== COMMIT FINAL ==========
        db.session.commit()

        # ========== PREPARAR RESPUESTA ==========
        detalles_response = []
        for detalle in detalles_creados:
            detalle_dict = detalle.to_dict()
            detalle_dict['producto_nombre'] = detalle.producto.nombre if detalle.producto else None
            detalle_dict['lote_codigo'] = detalle.lote.codigo_lote if detalle.lote else None
            detalles_response.append(detalle_dict)

        return created_response(
            data={
                'venta': nueva_venta.to_dict(),
                'detalles': detalles_response,
                'mensaje_exito': f'Venta {nueva_venta.numero_venta} procesada con exito'
            },
            message='Venta procesada exitosamente'
        )

    except ValueError as e:
        db.session.rollback()
        return validation_error_response(
            errors={'validacion': str(e)},
            message='Error de validacion en la venta'
        )
    except Exception as e:
        from flask import current_app
        import traceback
        current_app.logger.error(f"[VENTAS] Exception: {str(e)}")
        current_app.logger.error(f"[VENTAS] Traceback: {traceback.format_exc()}")
        db.session.rollback()
        return error_response(
            message='Error al procesar la venta',
            status_code=500,
            errors={'exception': str(e), 'traceback': traceback.format_exc()}
        )


# ==================================================================================
# ENDPOINT 2: GET /api/ventas - Listar ventas con filtros
# ==================================================================================

@ventas_bp.route('', methods=['GET'])
@jwt_required()
def listar_ventas():
    """
    Listar ventas con filtros opcionales

    Permite consultar el historial de ventas con diversos filtros:
    - Por fecha especifica
    - Por rango de fechas
    - Por vendedor
    - Por metodo de pago
    - Por estado

    Query Parameters:
        fecha (date): Filtrar por fecha especifica (YYYY-MM-DD)
        desde (date): Fecha inicio del rango
        hasta (date): Fecha fin del rango
        vendedor_id (int): Filtrar por vendedor
        metodo_pago (str): Filtrar por metodo de pago
        estado (str): Filtrar por estado
        limit (int): Maximo de resultados (default: 50, max: 200)
        offset (int): Paginacion (default: 0)

    Ordenamiento:
        Siempre por fecha DESC (mas reciente primero)

    Returns:
        200: Lista de ventas
        400: Parametros invalidos
        500: Error interno

    Ejemplo de respuesta:
        {
            "success": true,
            "message": "15 ventas encontradas",
            "data": {
                "ventas": [...],
                "total": 15,
                "limit": 50,
                "offset": 0
            }
        }
    """
    try:
        # ========== LOGS DE DEBUGGING ==========
        print(f'\n{"="*60}')
        print('[DEBUG] === GET /api/ventas ===')
        print(f'[DEBUG] Query params: {dict(request.args)}')

        # ========== CONSTRUIR QUERY BASE ==========
        query = Venta.query.filter(Venta.estado == 'completada')

        # ========== FILTRO POR FECHA ESPECIFICA ==========
        fecha_str = request.args.get('fecha')
        if fecha_str:
            print(f'[DEBUG] Filtrando por fecha específica: {fecha_str}')
            try:
                fecha_dt = datetime.strptime(fecha_str, '%Y-%m-%d').date()
                # IMPORTANTE: Usar timezone de Perú para comparar con datos timezone-aware
                fecha_inicio = datetime.combine(fecha_dt, datetime.min.time()).replace(tzinfo=PERU_TZ)
                fecha_fin = datetime.combine(fecha_dt, datetime.max.time()).replace(tzinfo=PERU_TZ)
                print(f'[DEBUG] Rango timezone-aware: {fecha_inicio} a {fecha_fin}')
                query = query.filter(
                    Venta.fecha >= fecha_inicio,
                    Venta.fecha <= fecha_fin
                )
            except ValueError:
                return error_response(
                    'Formato de fecha invalido. Use YYYY-MM-DD',
                    status_code=400
                )

        # ========== FILTRO POR RANGO DE FECHAS ==========
        # Soportar tanto 'desde/hasta' como 'fecha_inicio/fecha_fin'
        desde_str = request.args.get('desde') or request.args.get('fecha_inicio')
        hasta_str = request.args.get('hasta') or request.args.get('fecha_fin')

        if desde_str and hasta_str:
            print(f'[DEBUG] Filtrando por rango: {desde_str} a {hasta_str}')
            try:
                desde_dt = datetime.strptime(desde_str, '%Y-%m-%d').date()
                hasta_dt = datetime.strptime(hasta_str, '%Y-%m-%d').date()

                # IMPORTANTE: Usar timezone de Perú para comparar con datos timezone-aware
                desde_full = datetime.combine(desde_dt, datetime.min.time()).replace(tzinfo=PERU_TZ)
                hasta_full = datetime.combine(hasta_dt, datetime.max.time()).replace(tzinfo=PERU_TZ)

                print(f'[DEBUG] Rango timezone-aware: {desde_full} a {hasta_full}')

                query = query.filter(
                    Venta.fecha >= desde_full,
                    Venta.fecha <= hasta_full
                )
            except ValueError:
                return error_response(
                    'Formato de fecha invalido. Use YYYY-MM-DD',
                    status_code=400
                )

        # ========== FILTRO POR VENDEDOR ==========
        vendedor_id = request.args.get('vendedor_id', type=int)
        if vendedor_id:
            query = query.filter(Venta.vendedor_id == vendedor_id)

        # ========== FILTRO POR METODO DE PAGO ==========
        metodo_pago = request.args.get('metodo_pago')
        if metodo_pago:
            query = query.filter(Venta.metodo_pago == metodo_pago.lower())

        # ========== FILTRO POR ESTADO ==========
        estado = request.args.get('estado')
        if estado:
            query = query.filter(Venta.estado == estado.lower())

        # ========== ORDENAMIENTO ==========
        query = query.order_by(Venta.fecha.desc())

        # ========== PAGINACION ==========
        limit = request.args.get('limit', default=50, type=int)
        offset = request.args.get('offset', default=0, type=int)

        # Validar limit - permitir hasta 1000 ventas para filtros de fecha
        if limit > 1000:
            limit = 1000
        if limit < 1:
            limit = 50

        # Contar total antes de paginar
        total = query.count()
        print(f'\n[DEBUG] Total ventas encontradas: {total}')

        # Aplicar paginacion
        ventas = query.limit(limit).offset(offset).all()
        print(f'[DEBUG] Ventas retornadas (con limit={limit}): {len(ventas)}')

        # Mostrar las primeras ventas para debugging
        if ventas:
            print(f'\n[DEBUG] Primeras ventas encontradas:')
            for i, v in enumerate(ventas[:5], 1):
                print(f'  {i}. #{v.numero_venta} - Fecha: {v.fecha} - Total: S/ {v.total:.2f}')

        print(f'{"="*60}\n')

        # ========== ENRIQUECER DATOS ==========
        ventas_dict = []
        for idx, venta in enumerate(ventas):
            try:
                # CORRECCIÓN: Incluir detalles con precio_compra_unitario
                venta_data = venta.to_dict(include_detalles=True)
                # Agregar nombre del vendedor
                if venta.vendedor:
                    venta_data['vendedor_nombre'] = venta.vendedor.nombre_completo
                # Agregar cantidad de items
                venta_data['cantidad_items'] = len(venta.detalles) if venta.detalles else 0

                # Agregar alias para compatibilidad
                if 'fecha' in venta_data and venta_data['fecha']:
                    venta_data['fecha_venta'] = venta_data['fecha']

                ventas_dict.append(venta_data)
            except Exception as detalle_error:
                # Si falla una venta, loguear pero continuar con las demás
                print(f'[ERROR] Error al serializar venta {venta.id}: {str(detalle_error)}')
                import traceback
                traceback.print_exc()
                # Agregar versión simplificada sin detalles
                try:
                    venta_data_simple = venta.to_dict(include_detalles=False)
                    venta_data_simple['detalles'] = []
                    venta_data_simple['error'] = 'Error al cargar detalles'
                    ventas_dict.append(venta_data_simple)
                except:
                    pass

        # ========== RESPUESTA EXITOSA ==========
        return success_response(
            data={
                'ventas': ventas_dict,
                'total': total,
                'limit': limit,
                'offset': offset
            },
            message=f'{total} ventas encontradas'
        )

    except Exception as e:
        return error_response(
            message='Error al listar ventas',
            status_code=500,
            errors={'exception': str(e)}
        )


# ==================================================================================
# ENDPOINT 3: GET /api/ventas/<int:id> - Ver detalle de venta
# ==================================================================================

@ventas_bp.route('/<int:id>', methods=['GET'])
@jwt_required()
def obtener_venta(id):
    """
    Obtener detalle completo de una venta

    Retorna toda la informacion de la venta incluyendo:
    - Datos principales de la venta
    - Todos los items vendidos con info de producto y lote
    - Informacion del vendedor
    - Trazabilidad completa FIFO

    Path Parameter:
        id (int): ID de la venta

    Returns:
        200: Detalle de la venta
        404: Venta no encontrada
        500: Error interno

    Ejemplo de respuesta:
        {
            "success": true,
            "data": {
                "venta": {...},
                "detalles": [
                    {
                        "producto": {...},
                        "lote": {...},
                        "cantidad": 10,
                        "precio_unitario": 3.50,
                        "subtotal": 35.00,
                        "ganancia_unitaria": 1.20,
                        "ganancia_total": 12.00
                    }
                ],
                "vendedor": {...},
                "resumen": {
                    "cantidad_items": 2,
                    "cantidad_productos": 15,
                    "ganancia_total": 25.50
                }
            }
        }
    """
    try:
        # ========== BUSCAR LA VENTA ==========
        venta = Venta.query.get(id)
        if not venta:
            return not_found_response(f'Venta con ID {id} no encontrada')

        # ========== PREPARAR DETALLES CON INFORMACION COMPLETA ==========
        detalles_response = []
        cantidad_total_productos = 0
        ganancia_total = Decimal('0')

        for detalle in venta.detalles:
            detalle_dict = detalle.to_dict()

            # Agregar informacion del producto
            if detalle.producto:
                detalle_dict['producto'] = {
                    'id': detalle.producto.id,
                    'nombre': detalle.producto.nombre,
                    'codigo_barras': detalle.producto.codigo_barras,
                    'categoria': detalle.producto.categoria
                }

            # Agregar informacion del lote (trazabilidad)
            if detalle.lote:
                detalle_dict['lote'] = {
                    'id': detalle.lote.id,
                    'codigo_lote': detalle.lote.codigo_lote,
                    'fecha_vencimiento': detalle.lote.fecha_vencimiento.isoformat() if detalle.lote.fecha_vencimiento else None,
                    'proveedor': detalle.lote.proveedor
                }

            detalles_response.append(detalle_dict)

            # Calcular totales
            cantidad_total_productos += detalle.cantidad
            ganancia_total += detalle.ganancia_total

        # ========== PREPARAR INFORMACION DEL VENDEDOR ==========
        vendedor_dict = None
        if venta.vendedor:
            vendedor_dict = {
                'id': venta.vendedor.id,
                'nombre': venta.vendedor.nombre_completo,
                'email': venta.vendedor.email,
                'rol': venta.vendedor.rol
            }

        # ========== RESPUESTA EXITOSA ==========
        return success_response(
            data={
                'venta': venta.to_dict(),
                'detalles': detalles_response,
                'vendedor': vendedor_dict,
                'resumen': {
                    'cantidad_items': len(detalles_response),
                    'cantidad_productos': cantidad_total_productos,
                    'ganancia_total': str(ganancia_total)
                }
            },
            message='Detalle de venta obtenido exitosamente'
        )

    except Exception as e:
        return error_response(
            message='Error al obtener detalle de venta',
            status_code=500,
            errors={'exception': str(e)}
        )


# ==================================================================================
# ENDPOINT 4: GET /api/ventas/dia - Ventas del dia
# ==================================================================================

@ventas_bp.route('/dia', methods=['GET'])
@jwt_required()
def ventas_del_dia():
    """
    Obtener resumen de ventas del dia actual

    Retorna todas las ventas de hoy con estadisticas:
    - Lista completa de ventas
    - Total vendido en el dia
    - Cantidad de ventas
    - Desglose por metodo de pago
    - Ganancia total estimada

    Este endpoint es util para el cierre de caja diario.

    Returns:
        200: Resumen de ventas del dia
        500: Error interno

    Ejemplo de respuesta:
        {
            "success": true,
            "data": {
                "ventas": [...],
                "fecha": "2025-11-04",
                "total_dia": 450.50,
                "cantidad_ventas": 15,
                "ganancia_total": 125.30,
                "por_metodo_pago": {
                    "efectivo": 200.00,
                    "yape": 150.50,
                    "plin": 50.00,
                    "transferencia": 50.00
                },
                "por_estado": {
                    "completada": 14,
                    "cancelada": 1
                }
            }
        }
    """
    try:
        # ========== OBTENER VENTAS DEL DIA ==========
        hoy = date.today()
        ventas_hoy = Venta.ventas_del_dia()

        # ========== CALCULAR ESTADISTICAS ==========
        total_dia = Decimal('0')
        ganancia_total = Decimal('0')
        por_metodo_pago = {
            'efectivo': Decimal('0'),
            'yape': Decimal('0'),
            'plin': Decimal('0'),
            'transferencia': Decimal('0')
        }
        por_estado = {}

        ventas_dict = []
        for venta in ventas_hoy:
            venta_data = venta.to_dict()

            # Agregar nombre del vendedor
            if venta.vendedor:
                venta_data['vendedor_nombre'] = venta.vendedor.nombre_completo

            # Agregar cantidad de items
            venta_data['cantidad_items'] = len(venta.detalles)

            ventas_dict.append(venta_data)

            # Solo sumar si no esta cancelada
            if venta.estado != 'cancelada':
                total_dia += venta.total

                # Sumar por metodo de pago
                if venta.metodo_pago in por_metodo_pago:
                    por_metodo_pago[venta.metodo_pago] += venta.total

                # Calcular ganancia de esta venta
                ganancia_venta = venta.ganancia_total
                ganancia_total += ganancia_venta

            # Contar por estado
            if venta.estado in por_estado:
                por_estado[venta.estado] += 1
            else:
                por_estado[venta.estado] = 1

        # Convertir Decimal a string para JSON
        por_metodo_pago_str = {k: str(v) for k, v in por_metodo_pago.items()}

        # ========== RESPUESTA EXITOSA ==========
        return success_response(
            data={
                'ventas': ventas_dict,
                'fecha': hoy.isoformat(),
                'total_dia': str(total_dia),
                'cantidad_ventas': len(ventas_hoy),
                'ganancia_total': str(ganancia_total),
                'por_metodo_pago': por_metodo_pago_str,
                'por_estado': por_estado
            },
            message=f'{len(ventas_hoy)} ventas realizadas hoy'
        )

    except Exception as e:
        return error_response(
            message='Error al obtener ventas del dia',
            status_code=500,
            errors={'exception': str(e)}
        )


# ==================================================================================
# ENDPOINT 5: POST /api/ventas/<int:id>/cancelar - Cancelar venta
# ==================================================================================

@ventas_bp.route('/<int:id>/cancelar', methods=['POST'])
@jwt_required()
def cancelar_venta(id):
    """
    Cancelar una venta y revertir stock

    Marca la venta como cancelada y:
    1. Revierte el stock a los lotes originales
    2. Crea MovimientoStock tipo 'devolucion'
    3. Actualiza stock_total de productos
    4. Registra el motivo de cancelacion

    IMPORTANTE: La venta NO se elimina, solo se marca como cancelada
    para mantener trazabilidad y auditoria.

    Path Parameter:
        id (int): ID de la venta

    Body JSON:
        {
            "motivo": "Error en la venta"
        }

    Validaciones:
        - La venta debe existir
        - No puede estar ya cancelada
        - Motivo es requerido

    Returns:
        200: Venta cancelada exitosamente
        400: Venta ya esta cancelada
        404: Venta no encontrada
        422: Error de validacion
        500: Error interno
    """
    try:
        # ========== BUSCAR LA VENTA ==========
        venta = Venta.query.get(id)
        if not venta:
            return not_found_response(f'Venta con ID {id} no encontrada')

        # ========== VALIDAR QUE NO ESTE CANCELADA ==========
        if venta.estado == 'cancelada':
            return error_response(
                'La venta ya esta cancelada',
                status_code=400
            )

        # ========== VALIDAR MOTIVO ==========
        data = request.json
        motivo = data.get('motivo', '').strip() if data else ''

        if not motivo:
            return validation_error_response(
                errors={'motivo': 'El motivo de cancelacion es requerido'},
                message='Debe proporcionar un motivo de cancelacion'
            )

        # ========== REVERTIR STOCK DE CADA DETALLE ==========
        for detalle in venta.detalles:
            # Obtener producto y lote
            producto = detalle.producto
            lote = detalle.lote

            if not producto or not lote:
                continue  # Skip si no hay producto o lote

            # Guardar stock anterior
            stock_anterior = producto.stock_total

            # Devolver stock al lote
            lote.aumentar_stock(detalle.cantidad)

            # Actualizar stock del producto
            producto.stock_total += detalle.cantidad
            stock_nuevo = producto.stock_total

            # Crear MovimientoStock de devolucion
            movimiento = MovimientoStock(
                tipo='devolucion',
                producto_id=producto.id,
                lote_id=lote.id,
                usuario_id=venta.vendedor_id,
                venta_id=venta.id,
                cantidad=detalle.cantidad,  # Positivo porque es ingreso
                stock_anterior=stock_anterior,
                stock_nuevo=stock_nuevo,
                motivo=f'Cancelacion de venta {venta.numero_venta}',
                referencia=f'CANCEL-{venta.numero_venta}'
            )

            db.session.add(movimiento)

        # ========== MARCAR VENTA COMO CANCELADA ==========
        venta.estado = 'cancelada'

        # Agregar motivo a las notas
        nota_cancelacion = f'\n[CANCELADA] Motivo: {motivo}'
        if venta.notas:
            venta.notas += nota_cancelacion
        else:
            venta.notas = nota_cancelacion

        # ========== COMMIT ==========
        db.session.commit()

        # ========== RESPUESTA EXITOSA ==========
        return success_response(
            data={
                'venta': venta.to_dict(),
                'motivo_cancelacion': motivo
            },
            message='Venta cancelada exitosamente'
        )

    except ValueError as e:
        db.session.rollback()
        return error_response(
            message=str(e),
            status_code=400
        )
    except Exception as e:
        db.session.rollback()
        return error_response(
            message='Error al cancelar la venta',
            status_code=500,
            errors={'exception': str(e)}
        )


# ==================================================================================
# ENDPOINT 6: GET /api/ventas/reportes/resumen - Resumen de ventas
# ==================================================================================

@ventas_bp.route('/reportes/resumen', methods=['GET'])
@jwt_required()
def resumen_ventas():
    """
    Obtener resumen estadistico de ventas

    Genera un reporte con metricas clave de ventas:
    - Total vendido en el periodo
    - Cantidad de ventas
    - Ticket promedio
    - Producto mas vendido
    - Ganancia total estimada
    - Desglose por metodo de pago

    Query Parameters:
        desde (date): Fecha inicio (default: hoy)
        hasta (date): Fecha fin (default: hoy)

    Returns:
        200: Resumen de ventas
        400: Parametros invalidos
        500: Error interno

    Ejemplo de respuesta:
        {
            "success": true,
            "data": {
                "periodo": {
                    "desde": "2025-11-01",
                    "hasta": "2025-11-04"
                },
                "total_vendido": 1500.50,
                "cantidad_ventas": 45,
                "ticket_promedio": 33.34,
                "ganancia_total": 420.30,
                "producto_mas_vendido": {
                    "nombre": "Coca Cola 500ml",
                    "cantidad": 150
                },
                "por_metodo_pago": {...},
                "ventas_por_dia": {...}
            }
        }
    """
    try:
        # ========== OBTENER PARAMETROS ==========
        desde_str = request.args.get('fecha_inicio') or request.args.get('desde')
        hasta_str = request.args.get('fecha_fin') or request.args.get('hasta')

        # Defaults: hoy
        hoy = date.today()
        desde = hoy
        hasta = hoy

        if desde_str:
            try:
                desde = datetime.strptime(desde_str, '%Y-%m-%d').date()
            except ValueError:
                return error_response(
                    'Formato de fecha "desde" invalido. Use YYYY-MM-DD',
                    status_code=400
                )

        if hasta_str:
            try:
                hasta = datetime.strptime(hasta_str, '%Y-%m-%d').date()
            except ValueError:
                return error_response(
                    'Formato de fecha "hasta" invalido. Use YYYY-MM-DD',
                    status_code=400
                )

        # Validar que desde <= hasta
        if desde > hasta:
            return error_response(
                'La fecha "desde" debe ser anterior o igual a "hasta"',
                status_code=400
            )

        print(f'\n{"="*80}')
        print(f'📊 ENDPOINT /ventas/reportes/resumen')
        print(f'   📅 fecha_inicio recibida: "{desde_str}"')
        print(f'   📅 fecha_fin recibida: "{hasta_str}"')
        print(f'   📅 Fechas parseadas: {desde} a {hasta} (tipo: {type(desde).__name__})')

        # ========== CONSULTAR VENTAS DEL PERIODO ==========
        # SOLUCIÓN: Obtener TODAS las ventas y filtrar en Python por .date()
        # Esto evita problemas de zona horaria en SQL
        todas_ventas = Venta.query.filter(Venta.estado == 'completada').all()

        print(f'\n   📦 Total ventas en BD (completadas): {len(todas_ventas)}')

        # Filtrar con logs detallados
        ventas = []
        for v in todas_ventas:
            fecha_venta_date = v.fecha.date()
            incluir = desde <= fecha_venta_date <= hasta
            if incluir:
                ventas.append(v)
                print(f'      ✅ Venta #{v.numero_venta}: {v.fecha} -> date={fecha_venta_date} (INCLUIDA)')

        print(f'\n   ✅ Ventas en el rango [{desde}, {hasta}]: {len(ventas)}')

        # ========== CALCULAR METRICAS ==========
        total_vendido = Decimal('0')
        ganancia_total = Decimal('0')
        por_metodo_pago = {}
        por_vendedor = {}
        productos_vendidos = {}  # producto_id: cantidad

        print(f'\n💰 CALCULANDO GANANCIAS DEL RANGO {desde} a {hasta}:')
        print(f'💰 Cantidad de ventas a procesar: {len(ventas)}')
        print(f'{"─"*60}')

        for idx, venta in enumerate(ventas, 1):
            print(f'\n📝 VENTA #{idx}: {venta.numero_venta}')
            print(f'   ID: {venta.id}')
            print(f'   Fecha/Hora: {venta.fecha}')
            print(f'   Estado: {venta.estado}')

            total_vendido += venta.total
            ganancia_venta = venta.ganancia_total
            ganancia_total += ganancia_venta

            print(f'   Total: S/ {venta.total:.2f}')
            print(f'   Ganancia: S/ {ganancia_venta:.2f}')

            # Mostrar detalles de cada producto en la venta
            if venta.detalles:
                print(f'   Detalles ({len(venta.detalles)} productos):')
                for detalle in venta.detalles:
                    # Obtener nombre del producto a través de la relación
                    try:
                        nombre_producto = detalle.producto.nombre if detalle.producto else 'Producto desconocido'
                    except:
                        nombre_producto = 'Producto desconocido'

                    print(f'      • {nombre_producto}')
                    print(f'        - Precio venta: S/ {detalle.precio_unitario:.2f}')
                    print(f'        - Precio compra: S/ {detalle.precio_compra:.2f}')
                    print(f'        - Cantidad: {detalle.cantidad}')
                    print(f'        - Ganancia: (S/ {detalle.precio_unitario:.2f} - S/ {detalle.precio_compra:.2f}) × {detalle.cantidad} = S/ {detalle.ganancia_total:.2f}')

            # Contar por metodo de pago
            metodo = venta.metodo_pago or 'sin_especificar'
            if metodo in por_metodo_pago:
                por_metodo_pago[metodo]['cantidad'] += 1
                por_metodo_pago[metodo]['total'] += venta.total
            else:
                por_metodo_pago[metodo] = {
                    'metodo': metodo,
                    'cantidad': 1,
                    'total': venta.total
                }

            # Contar por vendedor
            vendedor_id = venta.vendedor_id
            if vendedor_id:
                if vendedor_id in por_vendedor:
                    por_vendedor[vendedor_id]['cantidad'] += 1
                    por_vendedor[vendedor_id]['total'] += venta.total
                else:
                    vendedor_nombre = venta.vendedor.nombre_completo if venta.vendedor else 'Sin nombre'
                    por_vendedor[vendedor_id] = {
                        'vendedor_id': vendedor_id,
                        'vendedor_nombre': vendedor_nombre,
                        'cantidad': 1,
                        'total': venta.total
                    }

            # Contar productos vendidos
            for detalle in venta.detalles:
                if detalle.producto_id in productos_vendidos:
                    productos_vendidos[detalle.producto_id]['cantidad'] += detalle.cantidad
                else:
                    producto_nombre = detalle.producto.nombre if detalle.producto else 'Desconocido'
                    productos_vendidos[detalle.producto_id] = {
                        'nombre': producto_nombre,
                        'cantidad': detalle.cantidad
                    }

        print(f'\n✅ RESULTADOS:')
        print(f'   Total vendido: S/ {total_vendido:.2f}')
        print(f'   Ganancia bruta: S/ {ganancia_total:.2f}')
        print(f'{"="*70}\n')

        # Calcular ticket promedio
        cantidad_ventas = len(ventas)
        ticket_promedio = (total_vendido / cantidad_ventas) if cantidad_ventas > 0 else Decimal('0')

        # Producto mas vendido
        producto_mas_vendido = None
        if productos_vendidos:
            producto_top = max(productos_vendidos.items(), key=lambda x: x[1]['cantidad'])
            producto_mas_vendido = {
                'producto_id': producto_top[0],
                'nombre': producto_top[1]['nombre'],
                'cantidad': producto_top[1]['cantidad']
            }

        # Convertir a listas para JSON
        ventas_por_metodo = [
            {
                'metodo': v['metodo'],
                'cantidad': v['cantidad'],
                'total': float(v['total'])
            }
            for v in por_metodo_pago.values()
        ]

        ventas_por_vendedor = [
            {
                'vendedor_id': v['vendedor_id'],
                'vendedor_nombre': v['vendedor_nombre'],
                'cantidad': v['cantidad'],
                'total': float(v['total'])
            }
            for v in por_vendedor.values()
        ]

        print(f'\n{"="*80}')
        print(f'✅ RESULTADOS FINALES:')
        print(f'   💰 Total vendido: S/ {total_vendido:.2f}')
        print(f'   💵 Ganancia bruta: S/ {ganancia_total:.2f}')
        print(f'   🛒 Cantidad de ventas: {len(ventas)}')
        print(f'   🎯 Ticket promedio: S/ {ticket_promedio:.2f}')
        print(f'{"="*80}\n')

        # ========== RESPUESTA EXITOSA ==========
        return success_response(
            data={
                'periodo': {
                    'desde': desde.isoformat(),
                    'hasta': hasta.isoformat()
                },
                'total_vendido': float(total_vendido),
                'cantidad_ventas': cantidad_ventas,
                'ticket_promedio': float(ticket_promedio),
                'ganancia_bruta': float(ganancia_total),  # CORRECCIÓN: Frontend espera ganancia_bruta
                'ganancia_total': float(ganancia_total),  # Mantener compatibilidad con SalesReport
                'producto_mas_vendido': producto_mas_vendido,
                'ventas_por_metodo': ventas_por_metodo,
                'ventas_por_vendedor': ventas_por_vendedor,
                'productos_vendidos': len(productos_vendidos)
            },
            message=f'Resumen de ventas del {desde} al {hasta}'
        )

    except Exception as e:
        import traceback
        print(f'❌ ERROR EN RESUMEN DE VENTAS: {str(e)}')
        print(f'❌ Traceback:')
        traceback.print_exc()
        return error_response(
            message='Error al generar resumen de ventas',
            status_code=500,
            errors={'exception': str(e), 'traceback': traceback.format_exc()}
        )


# ==================================================================================
# ENDPOINT: GET /api/ventas/reportes/pdf - Exportar reporte a PDF
# ==================================================================================

@ventas_bp.route('/reportes/pdf', methods=['GET'])
@jwt_required()
def exportar_reporte_pdf():
    """
    Genera un reporte de ventas profesional en formato PDF con Top 10 productos.

    Query Parameters:
        - fecha_inicio: Fecha inicio (YYYY-MM-DD)
        - fecha_fin: Fecha fin (YYYY-MM-DD)

    Returns:
        PDF file con el reporte de ventas
    """
    try:
        # Obtener parámetros
        fecha_inicio_str = request.args.get('fecha_inicio')
        fecha_fin_str = request.args.get('fecha_fin')

        # Defaults: hoy
        hoy = date.today()
        fecha_inicio = hoy
        fecha_fin = hoy

        if fecha_inicio_str:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        if fecha_fin_str:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()

        # Consultar ventas
        ventas = Venta.query.filter(
            and_(
                func.date(Venta.created_at) >= fecha_inicio,
                func.date(Venta.created_at) <= fecha_fin,
                Venta.estado == 'completada'
            )
        ).order_by(Venta.created_at.desc()).all()

        # Calcular totales
        total_vendido = sum(venta.total for venta in ventas)
        ganancia_total = sum(venta.ganancia_total for venta in ventas)
        total_unidades = sum(sum(detalle.cantidad for detalle in venta.detalles) for venta in ventas)

        # Calcular margen de ganancia
        margen_porcentaje = (ganancia_total / total_vendido * 100) if total_vendido > 0 else 0

        # Calcular Top 10 productos
        productos_ventas = {}
        for venta in ventas:
            for detalle in venta.detalles:
                producto_id = detalle.producto_id
                if producto_id not in productos_ventas:
                    productos_ventas[producto_id] = {
                        'nombre': detalle.producto.nombre if detalle.producto else 'Sin nombre',
                        'cantidad': 0,
                        'total': Decimal('0'),
                        'ganancia': Decimal('0')
                    }
                productos_ventas[producto_id]['cantidad'] += detalle.cantidad
                productos_ventas[producto_id]['total'] += detalle.subtotal
                productos_ventas[producto_id]['ganancia'] += detalle.ganancia_unitaria * detalle.cantidad

        top_productos = sorted(productos_ventas.values(), key=lambda x: x['total'], reverse=True)[:10]

        # Calcular estadísticas de métodos de pago
        metodos_count = {}
        for venta in ventas:
            metodo = venta.metodo_pago.upper()
            metodos_count[metodo] = metodos_count.get(metodo, 0) + 1
        metodo_mas_usado = max(metodos_count.items(), key=lambda x: x[1])[0] if metodos_count else 'N/A'

        # Calcular análisis por hora (hora pico)
        ventas_por_hora = {}
        for venta in ventas:
            hora = venta.created_at.hour
            ventas_por_hora[hora] = ventas_por_hora.get(hora, 0) + float(venta.total)

        hora_pico = 'N/A'
        if ventas_por_hora:
            hora_max = max(ventas_por_hora.items(), key=lambda x: x[1])[0]
            hora_pico = f'{hora_max}:00 - {hora_max+1}:00'

        # Comparación con período anterior
        dias_diferencia = (fecha_fin - fecha_inicio).days + 1
        fecha_inicio_anterior = fecha_inicio - timedelta(days=dias_diferencia)
        fecha_fin_anterior = fecha_inicio - timedelta(days=1)

        ventas_anterior = Venta.query.filter(
            and_(
                func.date(Venta.created_at) >= fecha_inicio_anterior,
                func.date(Venta.created_at) <= fecha_fin_anterior,
                Venta.estado == 'completada'
            )
        ).all()

        total_vendido_anterior = sum(venta.total for venta in ventas_anterior)
        comparacion = 'N/A'
        if total_vendido_anterior > 0:
            cambio_porcentaje = ((total_vendido - total_vendido_anterior) / total_vendido_anterior) * 100
            simbolo = '↑' if cambio_porcentaje > 0 else '↓'
            comparacion = f'{simbolo} {abs(cambio_porcentaje):.1f}% vs período anterior'
        elif len(ventas_anterior) == 0 and len(ventas) > 0:
            comparacion = '↑ Nuevo período (sin ventas previas)'

        # Crear PDF en memoria
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []

        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=22,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=20,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )

        # Título
        elements.append(Paragraph('KATITA POS - Reporte de Ventas', title_style))
        elements.append(Spacer(1, 0.15*inch))

        # Período
        periodo_text = f'Período: {fecha_inicio.strftime("%d/%m/%Y")} - {fecha_fin.strftime("%d/%m/%Y")}'
        elements.append(Paragraph(periodo_text, styles['Normal']))
        elements.append(Spacer(1, 0.25*inch))

        # Resumen de métricas
        resumen_data = [
            ['RESUMEN DE VENTAS', ''],
            ['Total de Ventas:', f'{len(ventas)} ventas'],
            ['Total Vendido:', f'S/ {total_vendido:.2f}'],
            ['Ganancia Total:', f'S/ {ganancia_total:.2f}'],
            ['Margen de Ganancia:', f'{margen_porcentaje:.1f}%'],
            ['Ticket Promedio:', f'S/ {(total_vendido / len(ventas)):.2f}' if len(ventas) > 0 else 'S/ 0.00'],
            ['Unidades Vendidas:', f'{total_unidades} unidades'],
            ['Método Más Usado:', metodo_mas_usado],
            ['Hora Pico:', hora_pico],
            ['Comparación:', comparacion]
        ]

        resumen_table = Table(resumen_data, colWidths=[3.2*inch, 2.5*inch])
        resumen_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 13),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ]))

        elements.append(resumen_table)
        elements.append(Spacer(1, 0.3*inch))

        # Top 10 Productos
        if top_productos:
            elements.append(Paragraph('TOP 10 PRODUCTOS MÁS VENDIDOS', styles['Heading2']))
            elements.append(Spacer(1, 0.15*inch))

            top_data = [['#', 'Producto', 'Cant.', 'Total Vendido', 'Ganancia']]
            for idx, prod in enumerate(top_productos, 1):
                top_data.append([
                    str(idx),
                    prod['nombre'][:30],
                    str(prod['cantidad']),
                    f"S/ {float(prod['total']):.2f}",
                    f"S/ {float(prod['ganancia']):.2f}"
                ])

            top_table = Table(top_data, colWidths=[0.4*inch, 2.5*inch, 0.8*inch, 1.2*inch, 1.2*inch])
            top_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#16a34a')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (1, 1), (1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgreen),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('TOPPADDING', (0, 1), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
            ]))

            elements.append(top_table)
            elements.append(Spacer(1, 0.3*inch))

        # Detalle de ventas
        if ventas:
            elements.append(Paragraph('DETALLE DE VENTAS', styles['Heading2']))
            elements.append(Spacer(1, 0.15*inch))

            ventas_data = [['Fecha', 'ID', 'Método', 'Vendedor', 'Total', 'Ganancia']]

            for venta in ventas:
                vendedor_nombre = venta.vendedor.nombre_completo if venta.vendedor else 'Sin asignar'
                ventas_data.append([
                    venta.created_at.strftime('%d/%m/%Y %H:%M'),
                    f'#{venta.id}',
                    venta.metodo_pago.upper(),
                    vendedor_nombre,
                    f'S/ {venta.total:.2f}',
                    f'S/ {venta.ganancia_total:.2f}'
                ])

            ventas_table = Table(ventas_data, colWidths=[1.2*inch, 0.5*inch, 1.1*inch, 1.5*inch, 0.9*inch, 0.9*inch])
            ventas_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('TOPPADDING', (0, 1), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ]))

            elements.append(ventas_table)

        # Generar PDF
        doc.build(elements)
        buffer.seek(0)

        # Nombre del archivo
        filename = f'reporte_ventas_{fecha_inicio}_{fecha_fin}.pdf'

        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return error_response(
            message='Error al generar PDF',
            status_code=500,
            errors={'exception': str(e)}
        )


# ==================================================================================
# ENDPOINT: GET /api/ventas/reportes/excel - Exportar reporte a Excel
# ==================================================================================

@ventas_bp.route('/reportes/excel', methods=['GET'])
@jwt_required()
def exportar_reporte_excel():
    """
    Genera un reporte de ventas profesional en formato Excel con Top 10 productos.

    Query Parameters:
        - fecha_inicio: Fecha inicio (YYYY-MM-DD)
        - fecha_fin: Fecha fin (YYYY-MM-DD)

    Returns:
        Excel file con el reporte de ventas
    """
    try:
        # Obtener parámetros
        fecha_inicio_str = request.args.get('fecha_inicio')
        fecha_fin_str = request.args.get('fecha_fin')

        # Defaults: hoy
        hoy = date.today()
        fecha_inicio = hoy
        fecha_fin = hoy

        if fecha_inicio_str:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        if fecha_fin_str:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()

        # Consultar ventas
        ventas = Venta.query.filter(
            and_(
                func.date(Venta.created_at) >= fecha_inicio,
                func.date(Venta.created_at) <= fecha_fin,
                Venta.estado == 'completada'
            )
        ).order_by(Venta.created_at.desc()).all()

        # Calcular totales
        total_vendido = sum(venta.total for venta in ventas)
        ganancia_total = sum(venta.ganancia_total for venta in ventas)
        total_unidades = sum(sum(detalle.cantidad for detalle in venta.detalles) for venta in ventas)

        # Calcular margen de ganancia
        margen_porcentaje = (ganancia_total / total_vendido * 100) if total_vendido > 0 else 0

        # Calcular Top 10 productos
        productos_ventas = {}
        for venta in ventas:
            for detalle in venta.detalles:
                producto_id = detalle.producto_id
                if producto_id not in productos_ventas:
                    productos_ventas[producto_id] = {
                        'nombre': detalle.producto.nombre if detalle.producto else 'Sin nombre',
                        'cantidad': 0,
                        'total': Decimal('0'),
                        'ganancia': Decimal('0')
                    }
                productos_ventas[producto_id]['cantidad'] += detalle.cantidad
                productos_ventas[producto_id]['total'] += detalle.subtotal
                productos_ventas[producto_id]['ganancia'] += detalle.ganancia_unitaria * detalle.cantidad

        top_productos = sorted(productos_ventas.values(), key=lambda x: x['total'], reverse=True)[:10]

        # Calcular estadísticas de métodos de pago
        metodos_count = {}
        for venta in ventas:
            metodo = venta.metodo_pago.upper()
            metodos_count[metodo] = metodos_count.get(metodo, 0) + 1
        metodo_mas_usado = max(metodos_count.items(), key=lambda x: x[1])[0] if metodos_count else 'N/A'

        # Calcular análisis por hora (hora pico)
        ventas_por_hora = {}
        for venta in ventas:
            hora = venta.created_at.hour
            ventas_por_hora[hora] = ventas_por_hora.get(hora, 0) + float(venta.total)

        hora_pico = 'N/A'
        if ventas_por_hora:
            hora_max = max(ventas_por_hora.items(), key=lambda x: x[1])[0]
            hora_pico = f'{hora_max}:00 - {hora_max+1}:00'

        # Comparación con período anterior
        dias_diferencia = (fecha_fin - fecha_inicio).days + 1
        fecha_inicio_anterior = fecha_inicio - timedelta(days=dias_diferencia)
        fecha_fin_anterior = fecha_inicio - timedelta(days=1)

        ventas_anterior = Venta.query.filter(
            and_(
                func.date(Venta.created_at) >= fecha_inicio_anterior,
                func.date(Venta.created_at) <= fecha_fin_anterior,
                Venta.estado == 'completada'
            )
        ).all()

        total_vendido_anterior = sum(venta.total for venta in ventas_anterior)
        comparacion = 'N/A'
        if total_vendido_anterior > 0:
            cambio_porcentaje = ((total_vendido - total_vendido_anterior) / total_vendido_anterior) * 100
            simbolo = '↑' if cambio_porcentaje > 0 else '↓'
            comparacion = f'{simbolo} {abs(cambio_porcentaje):.1f}% vs período anterior'
        elif len(ventas_anterior) == 0 and len(ventas) > 0:
            comparacion = '↑ Nuevo período (sin ventas previas)'

        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = 'Reporte de Ventas'

        # Estilos
        header_fill = PatternFill(start_color='1e40af', end_color='1e40af', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        title_font = Font(bold=True, size=16, color='1e40af')
        green_fill = PatternFill(start_color='16a34a', end_color='16a34a', fill_type='solid')
        green_header_font = Font(color='FFFFFF', bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Título
        ws.merge_cells('A1:F1')
        ws['A1'] = 'KATITA POS - Reporte de Ventas'
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25

        # Período
        ws.merge_cells('A2:F2')
        ws['A2'] = f'Período: {fecha_inicio.strftime("%d/%m/%Y")} - {fecha_fin.strftime("%d/%m/%Y")}'
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A2'].font = Font(size=11)

        # Resumen de métricas
        ws['A4'] = 'RESUMEN DE VENTAS'
        ws['A4'].font = Font(bold=True, size=13, color='1e40af')

        ws['A5'] = 'Total de Ventas:'
        ws['B5'] = f'{len(ventas)} ventas'
        ws['A6'] = 'Total Vendido:'
        ws['B6'] = f'S/ {total_vendido:.2f}'
        ws['A7'] = 'Ganancia Total:'
        ws['B7'] = f'S/ {ganancia_total:.2f}'
        ws['A8'] = 'Margen de Ganancia:'
        ws['B8'] = f'{margen_porcentaje:.1f}%'
        ws['A9'] = 'Ticket Promedio:'
        ws['B9'] = f'S/ {(total_vendido / len(ventas)):.2f}' if len(ventas) > 0 else 'S/ 0.00'
        ws['A10'] = 'Unidades Vendidas:'
        ws['B10'] = f'{total_unidades} unidades'
        ws['A11'] = 'Método Más Usado:'
        ws['B11'] = metodo_mas_usado
        ws['A12'] = 'Hora Pico:'
        ws['B12'] = hora_pico
        ws['A13'] = 'Comparación:'
        ws['B13'] = comparacion

        # Formatear resumen
        for row in range(5, 14):
            ws[f'A{row}'].font = Font(bold=True, size=10)
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            ws[f'B{row}'].alignment = Alignment(horizontal='left')

        # Top 10 Productos
        current_row = 15
        if top_productos:
            ws[f'A{current_row}'] = 'TOP 10 PRODUCTOS MÁS VENDIDOS'
            ws[f'A{current_row}'].font = Font(bold=True, size=13, color='16a34a')
            current_row += 2

            # Headers Top 10
            top_headers = ['#', 'Producto', 'Cantidad', 'Total Vendido', 'Ganancia']
            for col, header in enumerate(top_headers, start=1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = header
                cell.fill = green_fill
                cell.font = green_header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            # Datos Top 10
            current_row += 1
            for idx, prod in enumerate(top_productos, 1):
                ws.cell(row=current_row, column=1, value=idx)
                ws.cell(row=current_row, column=2, value=prod['nombre'])
                ws.cell(row=current_row, column=3, value=prod['cantidad'])
                ws.cell(row=current_row, column=4, value=f"S/ {float(prod['total']):.2f}")
                ws.cell(row=current_row, column=5, value=f"S/ {float(prod['ganancia']):.2f}")

                for col in range(1, 6):
                    ws.cell(row=current_row, column=col).border = border
                    ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='center' if col in [1, 3] else 'left')

                current_row += 1

            current_row += 1

        # Detalle de ventas
        ws[f'A{current_row}'] = 'DETALLE DE VENTAS'
        ws[f'A{current_row}'].font = Font(bold=True, size=13, color='1e40af')
        current_row += 2

        # Headers
        headers = ['Fecha', 'ID Venta', 'Método de Pago', 'Vendedor', 'Total', 'Ganancia']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Datos de ventas
        current_row += 1
        for venta in ventas:
            vendedor_nombre = venta.vendedor.nombre_completo if venta.vendedor else 'Sin asignar'
            ws.cell(row=current_row, column=1, value=venta.created_at.strftime('%d/%m/%Y %H:%M'))
            ws.cell(row=current_row, column=2, value=f'#{venta.id}')
            ws.cell(row=current_row, column=3, value=venta.metodo_pago.upper())
            ws.cell(row=current_row, column=4, value=vendedor_nombre)
            ws.cell(row=current_row, column=5, value=f'S/ {venta.total:.2f}')
            ws.cell(row=current_row, column=6, value=f'S/ {venta.ganancia_total:.2f}')

            # Aplicar bordes
            for col in range(1, 7):
                ws.cell(row=current_row, column=col).border = border
                ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='center')

            current_row += 1

        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15

        # Guardar en memoria
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Nombre del archivo
        filename = f'reporte_ventas_{fecha_inicio}_{fecha_fin}.xlsx'

        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return error_response(
            message='Error al generar Excel',
            status_code=500,
            errors={'exception': str(e)}
        )
